import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.urls import path
from django.http import HttpResponse
from django import forms
from .models import UserProfile, Journal, JournalLog, JournalScore, Role


# ── Form tambah user baru (dengan role & leader) ──────────────────────────────

class UserWithProfileCreationForm(UserCreationForm):
    first_name = forms.CharField(label='Nama Depan', max_length=150, required=True)
    last_name  = forms.CharField(label='Nama Belakang', max_length=150, required=False)
    email      = forms.EmailField(label='Email', required=False)
    role       = forms.ChoiceField(label='Role', choices=Role.choices, required=True)
    manager    = forms.ModelChoiceField(
        label='Leader (khusus Writer)',
        queryset=UserProfile.objects.filter(role=Role.MANAGER),
        required=False,
        help_text='Wajib diisi jika role = Writer',
    )

    class Meta(UserCreationForm.Meta):
        model = User
        fields = ('username', 'first_name', 'last_name', 'email', 'password1', 'password2')

    def clean(self):
        cleaned = super().clean()
        if cleaned.get('role') == Role.SUPERVISOR and not cleaned.get('manager'):
            self.add_error('manager', 'Leader wajib dipilih untuk role Writer.')
        return cleaned


class UserWithProfileChangeForm(UserChangeForm):
    role    = forms.ChoiceField(label='Role', choices=Role.choices, required=False)
    manager = forms.ModelChoiceField(
        label='Leader',
        queryset=UserProfile.objects.filter(role=Role.MANAGER),
        required=False,
    )

    class Meta(UserChangeForm.Meta):
        model = User

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            try:
                profile = self.instance.profile
                self.fields['role'].initial    = profile.role
                self.fields['manager'].initial = profile.manager
            except UserProfile.DoesNotExist:
                pass


# ── Custom UserAdmin ──────────────────────────────────────────────────────────

class UserWithProfileAdmin(BaseUserAdmin):
    add_form    = UserWithProfileCreationForm
    form        = UserWithProfileChangeForm
    list_display = ('username', 'get_full_name', 'email', 'get_role', 'get_manager', 'is_superuser')
    list_filter  = ('is_superuser', 'is_staff', 'profile__role')
    search_fields = ('username', 'first_name', 'last_name', 'email')

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': (
                'username', 'first_name', 'last_name', 'email',
                'password1', 'password2',
                'role', 'manager',
            ),
        }),
    )

    fieldsets = (
        (None,            {'fields': ('username', 'password')}),
        ('Informasi Pribadi', {'fields': ('first_name', 'last_name', 'email')}),
        ('Role Sistem',   {'fields': ('role', 'manager')}),
        ('Hak Akses',     {'fields': ('is_active', 'is_staff', 'is_superuser')}),
        ('Tanggal',       {'fields': ('last_login', 'date_joined')}),
    )

    def get_role(self, obj):
        try:
            return obj.profile.get_role_display()
        except UserProfile.DoesNotExist:
            return '-'
    get_role.short_description = 'Role'

    def get_manager(self, obj):
        try:
            m = obj.profile.manager
            return m.user.get_full_name() if m else '-'
        except UserProfile.DoesNotExist:
            return '-'
    get_manager.short_description = 'Leader'

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        role    = form.cleaned_data.get('role')
        manager = form.cleaned_data.get('manager')
        if role:
            profile, _ = UserProfile.objects.get_or_create(user=obj)
            profile.role    = role
            profile.manager = manager if role == Role.SUPERVISOR else None
            profile.save()

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view),
                 name='auth_user_import_excel'),
            path('import-excel/template/', self.admin_site.admin_view(self.download_template),
                 name='auth_user_excel_template'),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    # ── Download template Excel ───────────────────────────────────────────────

    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template User'

        headers = ['username', 'password', 'nama_depan', 'nama_belakang',
                   'email', 'role', 'username_leader']
        header_fill  = PatternFill('solid', fgColor='8B0000')
        header_font  = Font(bold=True, color='FFD700')
        header_align = Alignment(horizontal='center')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = 20

        # Baris contoh
        ws.append(['budi.santoso', 'Pass@123', 'Budi', 'Santoso',
                   'budi@contoh.com', 'supervisor', 'leader_username'])
        ws.append(['siti.rahayu', 'Pass@123', 'Siti', 'Rahayu',
                   'siti@contoh.com', 'manager', ''])

        # Sheet info role
        ws2 = wb.create_sheet('Info Role')
        ws2.append(['Kode Role', 'Keterangan'])
        for code, label in Role.choices:
            ws2.append([code, label])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="template_import_user.xlsx"'
        return response

    # ── Import Excel ─────────────────────────────────────────────────────────

    def import_excel_view(self, request):
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
            except Exception:
                messages.error(request, 'File tidak valid. Pastikan format .xlsx')
                return redirect('..')

            created = skipped = errors = 0
            error_list = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                username    = str(row[0] or '').strip()
                password    = str(row[1] or '').strip()
                first_name  = str(row[2] or '').strip()
                last_name   = str(row[3] or '').strip()
                email       = str(row[4] or '').strip()
                role        = str(row[5] or '').strip().lower()
                mgr_uname   = str(row[6] or '').strip()

                if not username or not password:
                    error_list.append(f'Baris {row_num}: username/password kosong.')
                    errors += 1
                    continue

                valid_roles = [r[0] for r in Role.choices]
                if role not in valid_roles:
                    error_list.append(
                        f'Baris {row_num} ({username}): role "{role}" tidak valid. '
                        f'Pilihan: {", ".join(valid_roles)}'
                    )
                    errors += 1
                    continue

                if User.objects.filter(username=username).exists():
                    skipped += 1
                    continue

                # Cari manager
                manager_profile = None
                if role == Role.SUPERVISOR and mgr_uname:
                    try:
                        mgr_user = User.objects.get(username=mgr_uname)
                        manager_profile = mgr_user.profile
                    except (User.DoesNotExist, UserProfile.DoesNotExist):
                        error_list.append(
                            f'Baris {row_num} ({username}): leader "{mgr_uname}" tidak ditemukan.'
                        )
                        errors += 1
                        continue

                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                )
                UserProfile.objects.create(
                    user=user,
                    role=role,
                    manager=manager_profile,
                )
                created += 1

            if created:
                messages.success(request, f'{created} user berhasil diimport.')
            if skipped:
                messages.warning(request, f'{skipped} user dilewati (username sudah ada).')
            if errors:
                messages.error(request, f'{errors} baris gagal:')
                for e in error_list[:10]:
                    messages.error(request, e)

            return redirect('../')

        return render(request, 'admin/import_excel.html', {
            'title': 'Import User dari Excel',
            'opts': self.model._meta,
        })


admin.site.unregister(User)
admin.site.register(User, UserWithProfileAdmin)


# ── Model lain ────────────────────────────────────────────────────────────────

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display  = ['user', 'role', 'manager']
    list_filter   = ['role']
    search_fields = ['user__username', 'user__first_name', 'user__last_name']


@admin.register(Journal)
class JournalAdmin(admin.ModelAdmin):
    list_display  = ['title', 'author', 'status', 'revision_count', 'published_at', 'updated_at']
    list_filter   = ['status']
    search_fields = ['title', 'author__user__username']


@admin.register(JournalScore)
class JournalScoreAdmin(admin.ModelAdmin):
    list_display = ['journal', 'scorer', 'total_score', 'recommendation', 'created_at']
    list_filter  = ['recommendation']


@admin.register(JournalLog)
class JournalLogAdmin(admin.ModelAdmin):
    list_display = ['journal', 'action', 'by_user', 'timestamp']
    list_filter  = ['action']
